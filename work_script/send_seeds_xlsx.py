import json
import requests

from openpyxl import load_workbook


ag_api = 'http://cain.search.dev.1datatech.cn/api/seed'            # 示例：生产search环境

def get_seed(val, desc):
    data = {
        "val": val,                              # 种子值
        "description": desc,                     # 种子描述
        "majorCategory": "seed_non_keyword",
        "minorCategory": "seed_forum_taizhou_common_thread_url",   # 种子类型
        "relationCode": "seed_forum_taizhou_common_thread_url",    # 种子类型
        "templateId": 583                        # 种子对应的模板ID
    }
    return data

headers = {
    'Content-Type': 'application/json;charset=UTF-8',
    'Authorization': 'whereareyou',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36',
}


def main():
    wb = load_workbook(filename='台州19楼-论坛列表采集需求.xlsx')
    ws = wb.active

    for row in range(4, ws.max_row + 1):
        seed_val = ws[f'd{row}'].value          # 种子值，a代表excel表中第一列（A）
        seed_desc = ws[f'e{row}'].value         # 种子描述，b代表excel表中第二列（B），实际操作前，可开启下面注释代码，进行验证。
        # 测试
        # print(f'val: {seed_val} desc: {seed_desc}')
        response = requests.post(ag_api, headers=headers, data=json.dumps(get_seed(seed_val,seed_desc)))
        print(response.text)

if __name__ == '__main__':
    main()