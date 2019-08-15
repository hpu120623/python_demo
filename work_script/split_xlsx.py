from openpyxl import Workbook
from openpyxl import load_workbook

"""
本脚本目的：
将xlsx文件的工作表按照指定行数进行拆分处理
"""


# 把一个数据集保存成一个xlsx文件，参数1：数据集，参数2：拆分文件的流水号。
def one_sheet(lines_1000, sheet_no):
    # 创建一个xlsx文件对象
    wb_ = Workbook()
    ws = wb_.active  # 取得默认的worksheet
    ws.title = '新测试表%02d' % (sheet_no + 1)  # 设置一个标题

    # 写头：循环写每个字段的值
    # TODO 如果没有表头则跳过该段代码
    header_idx = 0
    for col_ in first_line:
        ws.cell(row=1, column=(header_idx + 1)).value = col_.value
        header_idx += 1

    row_idx = 0  # 纪录行索引
    for row_ in lines_1000: # 遍历，读取传入的所有行数
        col_idx = 0
        for col_ in row_: # 遍历，写入每一行的数据
            ws.cell(row=(row_idx + 2), column=(col_idx + 1)).value = col_.value  # 有表头，数据从第2行开始
            col_idx += 1  # 纪录列索引
        row_idx += 1
    # 保存文件
    wb_.save(filename + '-%02d.xlsx' % (sheet_no + 1))


if __name__ == '__main__':
    filename = '对比缺失域名汇总'

    wb = load_workbook(filename + '.xlsx')
    ws = wb.active  # 默认获得第一个sheet
    # 得到所有worksheet
    sheet_names = wb.sheetnames  # 获得表单名字
    # 获取第一个worksheet
    sheet = wb[sheet_names[0]]
    # 把生成器转换为列表
    lines = list(sheet.rows)
    # 获取表头字段（默认只有一行）
    first_line = lines[0]
    # 获取数据行
    lines = lines[1:]

    # 限定每个文件中数据的行数
    batche_num = 400
    # 计算拆分文件个数
    batches = (len(lines) // batche_num) + 1
    # 循环写数据集到每一个文件
    for num in range(batches):
        # 取数据集，每个数据集最多1000行，最后一个不足1000行，直接处理。
        lines_ = lines[num * batche_num: (num + 1) * batche_num]
        one_sheet(lines_, num)
