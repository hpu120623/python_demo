import json
import time
import dateparser
import pymysql.cursors
import requests
import schedule
from urllib.parse import quote,unquote

from openpyxl import load_workbook
from retrying import retry

@retry(stop_max_attempt_number=3)
def get_ag_spider(request_api, page):
    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'Authorization': 'whereareyou',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36',
    }

    data = {"curPage": page,
            "pageSize": 100,
            "entity": {"name": '',
                       "alias": "",
                       "mediaId": "",
                       "kernelId": "",
                       "seedMajorCategory": "",
                       "seedMinorCategory": ""
                       }}
    response = requests.post(
        request_api,
        headers=headers,
        data=json.dumps(data),
        timeout=10
    )
    request_list = json.loads(response.text)['data']['list']
    if request_list:
        for spider in request_list:
            yield spider['name']


def get_kibana_spider():
    request_api = 'http://kibana.1datatech.cn/elasticsearch/_msearch'
    headers = {
        'Content-Type': 'application/x-ndjson',
        'Authorization': 'Basic d2ViOjFkYXRhJElORk8yMDE4MDc=',
        'kbn-xsrf': 'reporting',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36',
    }
    start_time = int(time.mktime(dateparser.parse('4小时前').timetuple())) * 1000
    end_time = int(time.time() * 1000)
    response = requests.post(
        request_api,
        headers=headers,
        data='\n{"index":["business-general-full-**"],"ignore_unavailable":true,"preference":%s}\n{"query":{"bool":{"must":[{"match_all":{}},{"range":{"fetchTime":{"gte":%s,"lte":%s,"format":"epoch_millis"}}}],"must_not":[]}},"size":0,"_source":{"excludes":[]},"aggs":{"2":{"terms":{"field":"origin.originName","size":10000,"order":{"_count":"desc"}}}}}\n' %
        (start_time,
         start_time,
         end_time),
        timeout=10)
    result_list = json.loads(response.text)[
        'responses'][0]['aggregations']['2']['buckets']
    for result in result_list:
        yield result['key']


# 获取空闲、跳转的爬虫name
# TODO 此处需要改善，目前人工确定
def get_idle_spider():
    wb = load_workbook(filename='idle_spider.xlsx')
    ws = wb.active

    idle_spider_set = set()
    for row in range(2, ws.max_row + 1):
        spider_name = ws[f'a{row}'].value
        spider_mark = ws[f'c{row}'].value
        if spider_mark == 2 or spider_mark == 0:
            idle_spider_set.add(spider_name)

    print(f'获取不会产生数据的爬虫,共计{len(idle_spider_set)}个')
    return idle_spider_set


# 获取两个集合的差集
def get_diff_set(set1, set2):
    return set1.difference(set2)

# 插入数据到mysql


def insert_data(data_set):
    connection = pymysql.connect(
        host='localhost',
        user='root',
        password='hpu123',
        db='compare_spider',
        charset='utf8',
        cursorclass=pymysql.cursors.DictCursor
    )
    try:
        with connection.cursor() as cursor:
            # Read a single record
            sql = "DELETE FROM `compare_spider`.`spider`"
            cursor.execute(sql)

        # connection is not autocommit by default. So you must commit to save
        # your changes.
        connection.commit()

        with connection.cursor() as cursor:
            for data in data_set:
                # Create a new record
                sql = "INSERT INTO `compare_spider`.`spider`(`name`) VALUES (%s)"
                cursor.execute(sql, data)

        connection.commit()
        print(f'缺失爬虫插入数据库成功,共计{len(data_set)}个')

    finally:
        connection.close()


def run():
    # meta和search请求api
    request_list = ['http://cain.1datatech.cn/api/origins',
                    'http://cain.search.1datatech.cn/api/origins']

    ag_spider_set = set()
    for request_api in request_list:
        for page in range(1, 3):
            for item in get_ag_spider(request_api, page):
                ag_spider_set.add(item)

    kibana_spider_set = set()
    for item in get_kibana_spider():
        kibana_spider_set.add(item)

    print(f'获取kibana产生数据的爬虫，共计{len(kibana_spider_set)}个')
    print(f'获取ag系统的爬虫，共计{len(ag_spider_set)}个')
    print(
        f'获取kibana缺失的爬虫，共计{len(get_diff_set(ag_spider_set, kibana_spider_set))}个')
    lost_kibana_set = get_diff_set(
        get_diff_set(
            ag_spider_set,
            kibana_spider_set),
        get_idle_spider())
    insert_data(lost_kibana_set)


if __name__ == '__main__':
    run()
    # schedule.every(5).minutes.do(run)
    # while True:
    #     schedule.run_pending()
    #     time.sleep(30)
    #     print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
