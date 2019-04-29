from pprint import pprint

from openpyxl import load_workbook

def open_yu():
    wb = load_workbook(filename='舆情通数据.xlsx')
    ws = wb.active
    news_dict = {}
    for row in range(2, ws.max_row + 1):
        yu_area = ws[f'a{row}'].value
        yu_url = ws[f'd{row}'].value
        yu_domain = ws[f'e{row}'].value
        # 创建字典
        if yu_domain in news_dict:
            news_dict[yu_domain].append(yu_url)
        else:
            news_dict[yu_domain] = [yu_area, yu_url]
    print('news_domain字典建立成功')
    return news_dict

def open_media(news_dict):
    wb = load_workbook(filename='media.xlsx')
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        news_domain = ws[f'c{row}'].value
        if news_domain in news_dict:
            print(f'第{row}行查找成功')
            ws[f'i{row}'].value = news_dict[news_domain][0]
            ws[f'j{row}'].value = news_dict[news_domain][1]
            if len(news_dict[news_domain][1:]) > 1:
                ws[f'k{row}'].value = f'共{len(news_dict[news_domain][1:])}个'

    wb.save('new_media.xlsx')

def main():
    """
    目的：本脚本用来获取两个excel表中数据对比，判断表1中数据是否存在表2，打标签
    总结：
        1.进行较大excel读取，使用openpyxl等库
        2.涉及到数据结构，优化时间复杂度，建立字典（字典结构的设计）
        3.可以考虑存储到内存redis中
    """
    open_media(open_yu())

if __name__ == '__main__':
    main()