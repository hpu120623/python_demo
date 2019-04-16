from openpyxl import load_workbook

def open_yu():
                wb = load_workbook(filename='舆情通数据.xlsx')
            ws = wb.active
            news_dict = {}
            for row in range(2, ws.max_row + 1):
                area_name = ws[f'i{row}'].value
                if not area_name:
