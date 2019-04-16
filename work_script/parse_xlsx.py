from openpyxl import load_workbook
from tld import get_fld

wb = load_workbook(filename='舆情通数据.xlsx')
ws = wb.active

for row in range(2, ws.max_row + 1):
    news_url = ws[f'd{row}'].value
    ws[f'e{row}'].value = get_fld(news_url)
    print(f'第{row}行:{news_url},domain是{get_fld(news_url)}')

wb.save('舆情通数据-域名.xlsx')


