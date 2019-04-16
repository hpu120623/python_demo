import redis
from openpyxl import load_workbook

def open_yu():
    wb = load_workbook(filename='舆情通数据.xlsx')
    ws = wb.active
    news_dict = {}
    for row in range(2, ws.max_row + 1):
        yu_area = ws[f'a{row}'].value
        yu_url = ws[f'd{row}'].value
        yu_domain = ws[f'e{row}'].value
        # 创建字典
        if yu_domain in news_dict:
            news_dict[yu_domain].append(yu_url)
        else:
            news_dict[yu_domain] = [yu_area, yu_url]
    print('news_domain字典建立成功')
    return news_dict

def con_redis(new_dict):
    pool = redis.ConnectionPool(host='127.0.0.1', port=6379, decode_responses=True)
    r = redis.Redis(connection_pool=pool)
    r.set('news_domain', new_dict)

if __name__ == '__main__':
    news_dict = open_yu()
    con_redis(news_dict)
