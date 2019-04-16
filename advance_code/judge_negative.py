# -*- coding: utf-8 -*-
# @Author  : Amos.Li
# @Time    : 2019/3/19 10:19
import json
import requests
from retrying import retry
from openpyxl import load_workbook


@retry(stop_max_attempt_number=3)
def negative_judge():

    NLP_API = 'http://nlp.beta.1datatech.cn/'

    headers = {
        'Content-Type': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36',
    }

    wb = load_workbook(filename='zonghe.xlsx')
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        data = {
            "title": ws[f'a{row}'].value,
            "type": ""
        }
        response = requests.post(url=NLP_API, data=json.dumps(data), headers=headers)
        result = json.loads(response.text)
        value = result.get('data', 0)
        if value == -1:
            ws[f'h{row}'].value = '负面'
        elif value == 0:
            ws[f'h{row}'].value = '中立'
        elif value == 1:
            ws[f'h{row}'].value = '正面'
    wb.save('今日头条App版综合正负面判断.xlsx')

negative_judge()